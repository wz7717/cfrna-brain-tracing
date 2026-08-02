#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import requests
from scipy.stats import hypergeom


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "manuscript/calculations/independent_enrichment"
SRC = OUT / "sources"
OUT.mkdir(parents=True, exist_ok=True)

PANEL_PATH = ROOT / "code/data/models/bo2023_saleem_network_top200_model_genes.csv"
PROJECTOR_PATH = (
    ROOT
    / "code/reproducibility/p0_bio3_projector/bo2023_reference_projector_linear_full.npz"
)
PRIMATE_XLSX = (
    SRC / "chiou2023_supp/tables/adh1914_Tables_S1_to_S19.xlsx"
)
HUMAN_XLSX = SRC / "siletti2023_cluster_annotation.xlsx"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def split_genes(value) -> set[str]:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return set()
    return {x.strip().upper() for x in str(value).split(",") if x.strip()}


def bh(pvalues: list[float]) -> np.ndarray:
    p = np.asarray(pvalues, dtype=float)
    order = np.argsort(p)
    out = np.ones(len(p))
    running = 1.0
    for rank in range(len(p) - 1, -1, -1):
        idx = order[rank]
        running = min(running, p[idx] * len(p) / (rank + 1))
        out[idx] = running
    return out


def enrichment_rows(
    marker_sets: dict[str, set[str]], panel: set[str], background: set[str], reference: str
) -> pd.DataFrame:
    rows = []
    for family, markers in marker_sets.items():
        eligible = markers & background
        overlap = panel & eligible
        M, n, N, k = len(background), len(eligible), len(panel & background), len(overlap)
        p = float(hypergeom.sf(k - 1, M, n, N)) if k else 1.0
        expected = N * n / M if M else np.nan
        fold = k / expected if expected else np.nan
        rows.append(
            {
                "reference": reference,
                "family": family,
                "background_n": M,
                "panel_n": N,
                "marker_n": n,
                "overlap_n": k,
                "expected_overlap": expected,
                "fold_enrichment": fold,
                "p_over": p,
                "overlap_genes": ";".join(sorted(overlap)),
            }
        )
    frame = pd.DataFrame(rows)
    frame["bh_q_7"] = bh(frame["p_over"].tolist())
    frame["significant_q05"] = frame["bh_q_7"] < 0.05
    return frame


def primate_markers() -> dict[str, set[str]]:
    workbook = openpyxl.load_workbook(PRIMATE_XLSX, read_only=True, data_only=True)
    class_sheet = workbook["S3"]
    subtype_sheet = workbook["S7"]
    by_class: dict[str, set[str]] = {}
    for row in class_sheet.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        by_class.setdefault(str(row[0]), set()).update(split_genes(row[5]))
    for row in subtype_sheet.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        by_class.setdefault(str(row[3]), set()).update(split_genes(row[10]))
    mapping = {
        "Excitatory neuron": ["glutamatergic neurons"],
        "Inhibitory neuron": [
            "GABAergic neurons",
            "basket cells",
            "medium spiny neurons",
        ],
        "Astrocyte": ["astrocytes"],
        "Oligodendrocyte": ["oligodendrocytes"],
        "OPC": ["oligodendrocyte precursor cells"],
        "Microglia": ["microglia", "KIR3DL12 microglia"],
        "Vascular": ["vascular cells"],
    }
    return {
        family: set().union(*(by_class.get(name, set()) for name in names))
        for family, names in mapping.items()
    }


def human_markers() -> dict[str, set[str]]:
    frame = pd.read_excel(HUMAN_XLSX)
    cls = frame["Class auto-annotation"].fillna("").astype(str)
    nt = frame["Neurotransmitter auto-annotation"].fillna("").astype(str)
    sc = frame["Supercluster"].fillna("").astype(str)
    masks = {
        "Excitatory neuron": cls.eq("NEUR") & nt.str.contains("VGLUT") & ~nt.str.contains("GABA"),
        "Inhibitory neuron": cls.eq("NEUR") & nt.str.contains("GABA"),
        "Astrocyte": cls.str.contains("ASTRO") | sc.eq("Astrocyte"),
        "Oligodendrocyte": sc.eq("Oligodendrocyte"),
        "OPC": sc.str.contains("oligodendrocyte precursor", case=False),
        "Microglia": sc.eq("Microglia") | cls.eq("MGL"),
        "Vascular": sc.eq("Vascular"),
    }
    result = {}
    for family, mask in masks.items():
        genes: set[str] = set()
        for value in frame.loc[mask, "Top Enriched Genes"]:
            genes.update(split_genes(value))
        result[family] = genes
    return result


def gprofiler(panel: set[str], background: set[str] | None, scope: str):
    payload = {
        "organism": "hsapiens",
        "query": sorted(panel),
        "sources": ["GO:BP", "KEGG"],
        "user_threshold": 0.05,
        "significance_threshold_method": "fdr",
        "no_evidences": False,
        "domain_scope": scope,
    }
    if background is not None:
        payload["background"] = sorted(background)
    response = requests.post(
        "https://biit.cs.ut.ee/gprofiler/api/gost/profile/",
        json=payload,
        timeout=180,
    )
    response.raise_for_status()
    return response.json()


def go_components(frame: pd.DataFrame, query_genes: list[str]) -> pd.DataFrame:
    go = frame.loc[frame["source"].eq("GO:BP")].copy()
    # With no_evidences=False, g:Profiler returns one evidence-code list per
    # successfully mapped query gene. Recover the intersecting gene symbols
    # from the non-empty positions before calculating term overlap.
    sets = [
        {gene for gene, evidence in zip(query_genes, values or []) if evidence}
        for values in go["intersections"]
    ]
    parent = list(range(len(go)))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for i in range(len(sets)):
        for j in range(i + 1, len(sets)):
            denominator = len(sets[i] | sets[j])
            if denominator and len(sets[i] & sets[j]) / denominator >= 0.5:
                union(i, j)
    go["component"] = [find(i) for i in range(len(go))]
    representatives = (
        go.sort_values(["p_value", "term_size"])
        .groupby("component", as_index=False)
        .first()
        .sort_values("p_value")
    )
    return representatives


panel_frame = pd.read_csv(PANEL_PATH)
panel = set(panel_frame["gene_symbol"].astype(str).str.upper())
background = set(
    np.load(PROJECTOR_PATH, allow_pickle=True)["genes"].astype(str)
)
if len(panel) != 200 or not panel <= background:
    raise RuntimeError(
        f"Panel/background mismatch: panel={len(panel)}, missing={len(panel-background)}"
    )

primate_sets = primate_markers()
human_sets = human_markers()
pd.DataFrame(
    [
        {"reference": "Chiou2023_rhesus", "family": k, "marker_gene": g}
        for k, values in primate_sets.items()
        for g in sorted(values)
    ]
).to_csv(OUT / "independent_primate_marker_sets.csv", index=False)
pd.DataFrame(
    [
        {"reference": "Siletti2023_human", "family": k, "marker_gene": g}
        for k, values in human_sets.items()
        for g in sorted(values)
    ]
).to_csv(OUT / "independent_human_marker_sets.csv", index=False)

cell_results = pd.concat(
    [
        enrichment_rows(primate_sets, panel, background, "Chiou2023_rhesus_primary"),
        enrichment_rows(human_sets, panel, background, "Siletti2023_human_sensitivity"),
    ],
    ignore_index=True,
)
cell_results.to_csv(OUT / "independent_celltype_enrichment.csv", index=False)

primary_json = gprofiler(panel, background, "custom")
sensitivity_json = gprofiler(panel, None, "annotated")
for name, payload in [
    ("gprofiler_model_background.json", primary_json),
    ("gprofiler_annotated_background.json", sensitivity_json),
]:
    (OUT / name).write_text(json.dumps(payload, indent=2), encoding="utf-8")

primary = pd.DataFrame(primary_json.get("result", []))
sensitivity = pd.DataFrame(sensitivity_json.get("result", []))
primary.to_csv(OUT / "gprofiler_model_background.csv", index=False)
sensitivity.to_csv(OUT / "gprofiler_annotated_background.csv", index=False)
query_genes = list(
    primary_json["meta"]["genes_metadata"]["query"]["query_1"]["mapping"].keys()
)
go_components(primary, query_genes).to_csv(
    OUT / "gprofiler_GO_BP_representative_components.csv", index=False
)

manifest = {
    "protocol": json.loads(
        (OUT / "analysis_protocol_20260731.json").read_text(encoding="utf-8")
    ),
    "input_sha256": {
        str(PANEL_PATH.relative_to(ROOT)): sha256(PANEL_PATH),
        str(PROJECTOR_PATH.relative_to(ROOT)): sha256(PROJECTOR_PATH),
        str(PRIMATE_XLSX.relative_to(ROOT)): sha256(PRIMATE_XLSX),
        str(HUMAN_XLSX.relative_to(ROOT)): sha256(HUMAN_XLSX),
    },
    "panel_n": len(panel),
    "background_n": len(background),
    "celltype_results": cell_results.to_dict(orient="records"),
    "gprofiler_primary_significant": {
        source: int((primary["source"] == source).sum())
        for source in sorted(primary["source"].unique())
    },
    "gprofiler_sensitivity_significant": {
        source: int((sensitivity["source"] == source).sum())
        for source in sorted(sensitivity["source"].unique())
    },
    "gprofiler_meta_primary": primary_json.get("meta", {}),
    "gprofiler_meta_sensitivity": sensitivity_json.get("meta", {}),
}
(OUT / "independent_enrichment_manifest.json").write_text(
    json.dumps(manifest, indent=2), encoding="utf-8"
)
print(json.dumps(manifest, indent=2))
